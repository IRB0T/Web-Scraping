from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
import openpyxl
import argparse


chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(executable_path='chromedriver.exe', options=chrome_options)
noelement = selenium.common.exceptions.NoSuchElementException

def get_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("-q", "--query", dest="query", help="Type your query")
    parser.add_argument("-l", "--location", dest="location", help="Enter Location to search")
    options = parser.parse_args()
    if not options.query:
        parser.error("[-] Please specify your query, use --help for more info")
    elif not options.location:
        parser.error("[-] Please specify location, use --help for more info")
    return options
options = get_arguments()
q = options.query
l = options.location

URL = "https://www.google.com/maps/search/" + q + " in " + l
driver.get(URL)

param = []
try:
    addiv = driver.find_elements_by_css_selector('div[data-result-ad-type]')
except noelement:
    addiv = 0
    pass
adcnt = len(addiv)
print("{0} ads detected".format(adcnt))
maindiv = driver.find_elements_by_css_selector('div.section-result-content')
print("Find {0} facilities.".format((len(maindiv) - adcnt)))
for i in range(adcnt, adcnt + 10):
    data = {}
    elem = maindiv[i]
    try:
        data['Potential Facility'] = elem.find_element_by_css_selector(
            'h3.section-result-title').text
        data['Type of Facility'] = elem.find_element_by_css_selector(
            'span.section-result-details').text
        data['Facility Address'] = elem.find_element_by_css_selector(
            'span.section-result-location').text
        data['Phone Number'] = elem.find_element_by_css_selector(
            'span.section-result-phone-number span').text
        data['Website Of Facility'] = elem.find_element_by_css_selector(
            'div.section-result-action-container a').get_attribute('href')
    except noelement:
        data['Type of Facility'] = "No data"
        data['Facility Address'] = "No data"
        data['Phone Number'] = "No data"
        data['Website Of Facility'] = "No data"
        pass
    param.append(data)

wb = openpyxl.Workbook()
wbstyle = openpyxl.styles.fonts
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Potential Facilities'
sheet.cell(row=1, column=2).value = 'Type Of Facility'
sheet.cell(row=1, column=3).value = 'Facility Address'
sheet.cell(row=1, column=4).value = 'Phone Number'
sheet.cell(row=1, column=5).value = 'Website Of Facility'
firstrow = sheet[1]
for cell in firstrow:
    cell.font = wbstyle.Font(bold=True, size=15)
i = 2
for x in param:
    sheet.cell(row=i, column=1).value = x['Potential Facility']
    sheet.cell(row=i, column=2).value = x['Type of Facility']
    sheet.cell(row=i, column=3).value = x['Facility Address']
    sheet.cell(row=i, column=4).value = x['Phone Number']
    sheet.cell(row=i, column=5).hyperlink = x['Website Of Facility']
    i += 1
print("Excel file created successfully.")
wb.save("gmap.xlsx")